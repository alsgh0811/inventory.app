from flask import Flask, render_template, redirect, request, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import openpyxl
import os
import csv

app = Flask(__name__)

# 🔐 운영용 SECRET KEY
app.secret_key = os.environ.get("SECRET_KEY")
if not app.secret_key:
    raise Exception("SECRET_KEY 환경변수가 설정되지 않았습니다.")

# 🗄 PostgreSQL만 사용
database_url = os.environ.get("DATABASE_URL")
if not database_url:
    raise Exception("DATABASE_URL 환경변수가 설정되지 않았습니다.")

if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ------------------ MODELS ------------------

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(50))
    is_active = db.Column(db.Boolean, default=True)
    is_superadmin = db.Column(db.Boolean, default=False)

    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=False)
    branch = db.relationship('Branch')


class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    spec = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    location = db.Column(db.String(100))

    histories = db.relationship("History", backref="item", cascade="all, delete-orphan")
    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=False)


class History(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey("item.id"))
    change_type = db.Column(db.String(10))
    quantity = db.Column(db.Integer)
    manager = db.Column(db.String(50))
    created_at = db.Column(
    db.DateTime,
    default=lambda: datetime.utcnow() + timedelta(hours=9)
)
    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=False)


class Branch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)

    def __repr__(self):
        return f"<Branch {self.name}>"


# ------------------ ROUTES ------------------

@app.route("/")
def home():
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])

    search = request.args.get("search")
    location_filter = request.args.get("location")

    query = Item.query.filter_by(branch_id=user.branch_id)

    if search:
        query = query.filter(
            Item.name.contains(search) |
            Item.spec.contains(search)
        )

    if location_filter:
        query = query.filter_by(location=location_filter)

    items = query.all()

    locations = db.session.query(Item.location)\
        .filter_by(branch_id=user.branch_id)\
        .distinct().all()

    locations = [l[0] for l in locations]

    return render_template(
        "index.html",
        items=items,
        locations=locations,
        selected_location=location_filter
    )


@app.route("/register", methods=["GET", "POST"])
def register():
    branches = Branch.query.all()

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        branch_id = request.form.get("branch_id")

        if not branch_id:
            return "사업소를 선택하세요."

        hashed_pw = generate_password_hash(password)

        new_user = User(
            username=username,
            password=hashed_pw,
            branch_id=int(branch_id),
            is_active=False   # ⭐ 승인 전까지 비활성
        )

        db.session.add(new_user)
        db.session.commit()

        return redirect("/login")

    return render_template("register.html", branches=branches)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = User.query.filter_by(username=request.form["username"]).first()

        if user and check_password_hash(user.password, request.form["password"]):

            if not user.is_active:
                return "아직 승인되지 않은 계정입니다."

            session["user_id"] = user.id
            return redirect("/")

        return "로그인 실패"

    return render_template("login.html")


@app.route("/admin")
def admin_page():

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])

    if not current_user.is_superadmin:
        return "접근 권한 없음"

    branch_id = request.args.get("branch_id")

    if branch_id:
        users = User.query.filter_by(branch_id=branch_id).all()
    else:
        users = User.query.all()

    branches = Branch.query.all()

    return render_template("admin.html",
                           users=users,
                           branches=branches,
                           selected_branch=branch_id)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/edit_user/<int:user_id>", methods=["GET", "POST"])
def edit_user(user_id):

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])
    if not current_user.is_superadmin:
        return "접근 권한 없음"

    user = User.query.get(user_id)
    branches = Branch.query.all()

    if request.method == "POST":
        user.branch_id = request.form["branch_id"]
        user.is_active = True if request.form.get("is_active") == "on" else False
        user.is_superadmin = True if request.form.get("is_superadmin") == "on" else False
        db.session.commit()
        return redirect("/admin")

    return render_template("edit_user.html", user=user, branches=branches)

@app.route("/delete_user/<int:user_id>")
def delete_user(user_id):

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])
    if not current_user.is_superadmin:
        return "접근 권한 없음"

    user = User.query.get(user_id)

    if user:
        db.session.delete(user)
        db.session.commit()

    return redirect("/admin")


@app.route("/approve/<int:user_id>")
def approve(user_id):

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])

    if not current_user.is_superadmin:
        return "접근 권한 없음"

    user = User.query.get_or_404(user_id)

    user.is_active = True
    db.session.commit()

    return redirect("/admin")


@app.route("/add_item", methods=["POST"])
def add_item():
    if "user_id" not in session:
        return redirect("/login")

    name = request.form["name"].strip()
    spec = request.form["spec"].strip()
    location = request.form["location"].strip()

    # ✅ 필수값 체크
    if not name or not spec or not location:
        return "이름, 규격, 위치는 필수입니다."

    try:
        quantity = int(request.form["quantity"])
    except:
        return "수량은 숫자만 입력하세요."

    user = User.query.get(session["user_id"])

    item = Item(
        name=name,
        spec=spec,
        quantity=quantity,
        location=location,
        branch_id=user.branch_id  # ⭐ 매우 중요
    )

    db.session.add(item)
    db.session.commit()

    return redirect("/")

@app.route("/edit_item/<int:item_id>", methods=["GET", "POST"])
def edit_item(item_id):
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])

    item = Item.query.filter_by(
        id=item_id,
        branch_id=user.branch_id
    ).first_or_404()

    if request.method == "POST":
        name = request.form.get("name")
        spec = request.form.get("spec")
        location = request.form.get("location")

        if not name or not spec or not location:
            return "모든 항목을 입력해야 합니다."

        item.name = name
        item.spec = spec
        item.location = location

        db.session.commit()
        return redirect("/")

    return render_template("edit_item.html", item=item)

@app.route("/delete_item/<int:item_id>")
def delete_item(item_id):
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])

    item = Item.query.filter_by(
        id=item_id,
        branch_id=user.branch_id
    ).first_or_404()

    db.session.delete(item)
    db.session.commit()

    return redirect("/")


@app.route("/update_stock/<int:item_id>", methods=["POST"])
def update_stock(item_id):
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])

    item = Item.query.filter_by(
        id=item_id,
        branch_id=user.branch_id
    ).first_or_404()

    change_type = request.form["type"]
    quantity = int(request.form["quantity"])

    if change_type == "IN":
        item.quantity += quantity

    elif change_type == "OUT":
        # 🔥 재고 부족 체크
        if item.quantity < quantity:
            return "재고가 부족합니다."

        item.quantity -= quantity

    # 히스토리 저장
    
    user = User.query.get(session["user_id"])

    history = History(
        item_id=item.id,
        change_type=change_type,
        quantity=quantity,
        branch_id=user.branch_id   # 🔥 추가
    )
    db.session.add(history)

    db.session.commit()

    return redirect("/")


@app.route("/history")
def history():
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])
    records = History.query.filter_by(
        branch_id=user.branch_id
    ).order_by(History.created_at.desc()).all()

    return render_template("history.html", records=records)

@app.route("/download_excel")
def download_excel():
    records = History.query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["자재명", "구분", "수량", "날짜"])

    for r in records:
        ws.append([
            r.item.name if r.item else "",
            r.change_type,
            r.quantity,
            r.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    file_path = "history.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

@app.route("/upload_csv", methods=["POST"])
def upload_csv():
    if "user_id" not in session:
        return redirect("/login")

    file = request.files["file"]

    if not file or file.filename == "":
        return "파일이 없습니다."

    # 🔥 utf-8-sig 로 BOM 제거
    stream = file.stream.read().decode("utf-8-sig").splitlines()
    csv_reader = csv.DictReader(stream)

    user = User.query.get(session["user_id"])

    for row in csv_reader:
        try:
            name = row["name"].strip()
            spec = row["spec"].strip()
            quantity = int(row["quantity"])
            location = row["location"].strip()

            if not name or not spec or not location:
                continue

            existing_item = Item.query.filter_by(
                name=name,
                spec=spec,
                branch_id=user.branch_id
            ).first()

            if existing_item:
                existing_item.quantity += quantity
            else:
                new_item = Item(
                    name=name,
                    spec=spec,
                    quantity=quantity,
                    location=location,
                    branch_id=user.branch_id
                )
                db.session.add(new_item)

        except Exception as e:
            print("CSV 오류:", e)
            continue

    db.session.commit()
    return redirect("/")

@app.route("/branches", methods=["GET", "POST"])
def manage_branches():

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])

    if not current_user.is_superadmin:
        return "접근 권한 없음"

    if request.method == "POST":
        branch_name = request.form["branch_name"]

        new_branch = Branch(name=branch_name)
        db.session.add(new_branch)
        db.session.commit()

        return redirect("/branches")

    branches = Branch.query.all()

    return render_template("branches.html", branches=branches)

@app.route("/delete_branch/<int:branch_id>")
def delete_branch(branch_id):

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])

    if not current_user.is_superadmin:
        return "접근 권한 없음"

    branch = Branch.query.get(branch_id)

    # 🔴 해당 사업소에 소속된 사용자 있는지 확인
    users_in_branch = User.query.filter_by(branch_id=branch_id).count()

    if users_in_branch > 0:
        return "해당 사업소에 소속된 사용자가 있어 삭제할 수 없습니다."

    db.session.delete(branch)
    db.session.commit()

    return redirect("/branches")


@app.context_processor
def inject_user():
    if "user_id" in session:
        user = User.query.get(session["user_id"])
        return dict(current_user=user)
    return dict(current_user=None)


from flask import Flask, render_template, redirect, request, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import openpyxl
import os
import csv


app = Flask(__name__)
app.secret_key = "supersecretkey"

import os

database_url = os.environ.get("DATABASE_URL")

if database_url:
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///instance/inventory.db'
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ------------------ MODELS ------------------

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(50))
    is_active = db.Column(db.Boolean, default=True)
    is_superadmin = db.Column(db.Boolean, default=False)

    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=False)
    branch = db.relationship('Branch')


class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    spec = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    location = db.Column(db.String(100))

    histories = db.relationship("History", backref="item", cascade="all, delete-orphan")
    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=False)


class History(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey("item.id"))
    change_type = db.Column(db.String(10))
    quantity = db.Column(db.Integer)
    manager = db.Column(db.String(50))
    created_at = db.Column(
    db.DateTime,
    default=lambda: datetime.utcnow() + timedelta(hours=9)
)
    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=False)


class Branch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)

    def __repr__(self):
        return f"<Branch {self.name}>"


# ------------------ ROUTES ------------------

@app.route("/")
def home():
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])

    search = request.args.get("search")
    location_filter = request.args.get("location")

    query = Item.query.filter_by(branch_id=user.branch_id)

    if search:
        query = query.filter(
            Item.name.contains(search) |
            Item.spec.contains(search)
        )

    if location_filter:
        query = query.filter_by(location=location_filter)

    items = query.all()

    locations = db.session.query(Item.location)\
        .filter_by(branch_id=user.branch_id)\
        .distinct().all()

    locations = [l[0] for l in locations]

    return render_template(
        "index.html",
        items=items,
        locations=locations,
        selected_location=location_filter
    )


@app.route("/register", methods=["GET", "POST"])
def register():
    branches = Branch.query.all()

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        branch_id = request.form.get("branch_id")

        if not branch_id:
            return "사업소를 선택하세요."

        hashed_pw = generate_password_hash(password)

        new_user = User(
            username=username,
            password=hashed_pw,
            branch_id=int(branch_id),
            is_active=False   # ⭐ 승인 전까지 비활성
        )

        db.session.add(new_user)
        db.session.commit()

        return redirect("/login")

    return render_template("register.html", branches=branches)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = User.query.filter_by(username=request.form["username"]).first()

        if user and check_password_hash(user.password, request.form["password"]):

            if not user.is_active:
                return "아직 승인되지 않은 계정입니다."

            session["user_id"] = user.id
            return redirect("/")

        return "로그인 실패"

    return render_template("login.html")


@app.route("/admin")
def admin_page():

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])

    if not current_user.is_superadmin:
        return "접근 권한 없음"

    branch_id = request.args.get("branch_id")

    if branch_id:
        users = User.query.filter_by(branch_id=branch_id).all()
    else:
        users = User.query.all()

    branches = Branch.query.all()

    return render_template("admin.html",
                           users=users,
                           branches=branches,
                           selected_branch=branch_id)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/edit_user/<int:user_id>", methods=["GET", "POST"])
def edit_user(user_id):

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])
    if not current_user.is_superadmin:
        return "접근 권한 없음"

    user = User.query.get(user_id)
    branches = Branch.query.all()

    if request.method == "POST":
        user.branch_id = request.form["branch_id"]
        user.is_active = True if request.form.get("is_active") == "on" else False
        user.is_superadmin = True if request.form.get("is_superadmin") == "on" else False
        db.session.commit()
        return redirect("/admin")

    return render_template("edit_user.html", user=user, branches=branches)

@app.route("/delete_user/<int:user_id>")
def delete_user(user_id):

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])
    if not current_user.is_superadmin:
        return "접근 권한 없음"

    user = User.query.get(user_id)

    if user:
        db.session.delete(user)
        db.session.commit()

    return redirect("/admin")


@app.route("/approve/<int:user_id>")
def approve(user_id):

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])

    if not current_user.is_superadmin:
        return "접근 권한 없음"

    user = User.query.get_or_404(user_id)

    user.is_active = True
    db.session.commit()

    return redirect("/admin")


@app.route("/add_item", methods=["POST"])
def add_item():
    if "user_id" not in session:
        return redirect("/login")

    name = request.form["name"].strip()
    spec = request.form["spec"].strip()
    location = request.form["location"].strip()

    # ✅ 필수값 체크
    if not name or not spec or not location:
        return "이름, 규격, 위치는 필수입니다."

    try:
        quantity = int(request.form["quantity"])
    except:
        return "수량은 숫자만 입력하세요."

    user = User.query.get(session["user_id"])

    item = Item(
        name=name,
        spec=spec,
        quantity=quantity,
        location=location,
        branch_id=user.branch_id  # ⭐ 매우 중요
    )

    db.session.add(item)
    db.session.commit()

    return redirect("/")

@app.route("/edit_item/<int:item_id>", methods=["GET", "POST"])
def edit_item(item_id):
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])

    item = Item.query.filter_by(
        id=item_id,
        branch_id=user.branch_id
    ).first_or_404()

    if request.method == "POST":
        name = request.form.get("name")
        spec = request.form.get("spec")
        location = request.form.get("location")

        if not name or not spec or not location:
            return "모든 항목을 입력해야 합니다."

        item.name = name
        item.spec = spec
        item.location = location

        db.session.commit()
        return redirect("/")

    return render_template("edit_item.html", item=item)

@app.route("/delete_item/<int:item_id>")
def delete_item(item_id):
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])

    item = Item.query.filter_by(
        id=item_id,
        branch_id=user.branch_id
    ).first_or_404()

    db.session.delete(item)
    db.session.commit()

    return redirect("/")


@app.route("/update_stock/<int:item_id>", methods=["POST"])
def update_stock(item_id):
    if "user_id" not in session:
        return redirect("/login")

    item = Item.query.get_or_404(item_id)

    change_type = request.form["type"]
    quantity = int(request.form["quantity"])

    if change_type == "IN":
        item.quantity += quantity

    elif change_type == "OUT":
        # 🔥 재고 부족 체크
        if item.quantity < quantity:
            return "재고가 부족합니다."

        item.quantity -= quantity

    # 히스토리 저장
    
    user = User.query.get(session["user_id"])

    history = History(
        item_id=item.id,
        change_type=change_type,
        quantity=quantity,
        branch_id=user.branch_id   # 🔥 추가
    )
    db.session.add(history)

    db.session.commit()

    return redirect("/")


@app.route("/history")
def history():
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])
    records = History.query.filter_by(
        branch_id=user.branch_id
    ).order_by(History.created_at.desc()).all()

    return render_template("history.html", records=records)

@app.route("/download_excel")
def download_excel():
    records = History.query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["자재명", "구분", "수량", "날짜"])

    for r in records:
        ws.append([
            r.item.name if r.item else "",
            r.change_type,
            r.quantity,
            r.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    file_path = "history.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

@app.route("/download_inventory")
def download_inventory():
    if "user_id" not in session:
        return redirect("/login")

    user = User.query.get(session["user_id"])

    # 현재 사용자의 사업소 재고만 가져오기
    items = Item.query.filter_by(branch_id=user.branch_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "현재재고"

    # 헤더
    ws.append(["이름", "규격", "수량", "위치"])

    # 데이터 입력
    for item in items:
        ws.append([
            item.name,
            item.spec,
            item.quantity,
            item.location
        ])

    filename = "inventory.xlsx"
    wb.save(filename)

    return send_file(filename, as_attachment=True)


@app.route("/upload_csv", methods=["POST"])
def upload_csv():

    if "file" not in request.files:
        return "파일이 없습니다."

    file = request.files["file"]

    if file.filename == "":
        return "파일이 선택되지 않았습니다."

    try:
        stream = io.StringIO(file.stream.read().decode("utf-8"))
        csv_reader = csv.DictReader(stream)

        for row in csv_reader:

            name = row["이름"]
            spec = row["규격"]
            qty = int(row["수량"])
            location = row["위치"]

            item = Item(
                name=name,
                spec=spec,
                quantity=qty,
                location=location
            )

            db.session.add(item)

        db.session.commit()

        return redirect("/")

    except Exception as e:
        return f"CSV 업로드 오류: {str(e)}"

@app.route("/branches", methods=["GET", "POST"])
def manage_branches():

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])

    if not current_user.is_superadmin:
        return "접근 권한 없음"

    if request.method == "POST":
        branch_name = request.form["branch_name"]

        new_branch = Branch(name=branch_name)
        db.session.add(new_branch)
        db.session.commit()

        return redirect("/branches")

    branches = Branch.query.all()

    return render_template("branches.html", branches=branches)

@app.route("/delete_branch/<int:branch_id>")
def delete_branch(branch_id):

    if "user_id" not in session:
        return redirect("/login")

    current_user = User.query.get(session["user_id"])

    if not current_user.is_superadmin:
        return "접근 권한 없음"

    branch = Branch.query.get(branch_id)

    # 🔴 해당 사업소에 소속된 사용자 있는지 확인
    users_in_branch = User.query.filter_by(branch_id=branch_id).count()

    if users_in_branch > 0:
        return "해당 사업소에 소속된 사용자가 있어 삭제할 수 없습니다."

    db.session.delete(branch)
    db.session.commit()

    return redirect("/branches")


@app.context_processor
def inject_user():
    if "user_id" in session:
        user = User.query.get(session["user_id"])
        return dict(current_user=user)
    return dict(current_user=None)


with app.app_context():
    db.create_all()

# ==============================
# 🔥 운영용 자동 DB 세팅 코드
# ==============================

from werkzeug.security import generate_password_hash

with app.app_context():
    db.create_all()

    try:
        admin = User.query.filter_by(username="admin").first()
        if not admin:

            # 🔥 먼저 기본 지점 하나 만들기 (없으면 생성)
            branch = Branch.query.first()
            if not branch:
                branch = Branch(name="본사")
                db.session.add(branch)
                db.session.commit()

            # 🔥 관리자 생성
            admin = User(
                username="admin",
                password=generate_password_hash("1234"),
                role="superadmin",
                is_superadmin=True,
                branch_id=branch.id
            )

            db.session.add(admin)
            db.session.commit()
            print("✅ 관리자 계정 생성 완료")

    except Exception as e:
        print("❌ 관리자 생성 오류:", e)


if __name__ == "__main__":
    app.run()
